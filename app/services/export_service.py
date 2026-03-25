"""
Servicios de exportación a CSV, PDF y Excel
"""
import csv
import io
from datetime import datetime
from flask import current_app
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import json
from app.models import AuditLog

class ExportService:
    """Servicio para exportar correos a diferentes formatos"""
    
    @staticmethod
    def to_csv(emails):
        """
        Exporta correos a formato CSV
        
        Args:
            emails (list): Lista de objetos Email
            
        Returns:
            bytes: Contenido CSV como bytes
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Encabezados
        headers = [
            'ID',
            'Fecha de Envío',
            'Asunto',
            'Remitente',
            'Destinatarios',
            'Importancia',
            'Archivos Adjuntos',
            'Fecha de Creación en Sistema'
        ]
        writer.writerow(headers)
        
        # Datos
        for email in emails:
            try:
                recipients = json.loads(email.recipients)
                recipient_emails = [r['email'] for r in recipients]
                recipient_str = '; '.join(recipient_emails)
            except:
                recipient_str = email.recipients
            
            writer.writerow([
                email.id,
                email.sent_date.strftime('%Y-%m-%d %H:%M:%S'),
                email.subject,
                email.sender,
                recipient_str,
                email.importance,
                'Sí' if email.has_attachments else 'No',
                email.created_at.strftime('%Y-%m-%d %H:%M:%S')
            ])
        
        return output.getvalue().encode('utf-8')
    
    @staticmethod
    def to_pdf(emails, title="Reporte de Correos Electrónicos"):
        """
        Exporta correos a formato PDF
        
        Args:
            emails (list): Lista de objetos Email
            title (str): Título del reporte
            
        Returns:
            bytes: Contenido PDF como bytes
        """
        pdf_buffer = io.BytesIO()
        
        # Crear documento PDF
        doc = SimpleDocTemplate(
            pdf_buffer,
            pagesize=letter,
            rightMargin=0.5*inch,
            leftMargin=0.5*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=12,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=10,
            textColor=colors.white,
            backColor=colors.HexColor('#1f77b4'),
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        cell_style = ParagraphStyle(
            'CustomCell',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_LEFT
        )
        
        # Contenido
        story = []
        
        # Título
        story.append(Paragraph(title, title_style))
        story.append(Paragraph(
            f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
            styles['Normal']
        ))
        story.append(Spacer(1, 12))
        
        # Tabla de correos
        if emails:
            data = [['ID', 'Fecha', 'Asunto', 'Remitente', 'Destinatarios', 'Adj.']]
            
            for email in emails:
                try:
                    recipients = json.loads(email.recipients)
                    recipient_emails = [r['email'] for r in recipients]
                    recipient_str = '; '.join(recipient_emails[:2])  # Primeros 2
                    if len(recipient_emails) > 2:
                        recipient_str += f" +{len(recipient_emails)-2}"
                except:
                    recipient_str = "Error"
                
                subject_short = email.subject[:30] + "..." if len(email.subject) > 30 else email.subject
                
                data.append([
                    str(email.id),
                    email.sent_date.strftime('%d/%m/%y'),
                    subject_short,
                    email.sender[:20] + "..." if len(email.sender) > 20 else email.sender,
                    recipient_str[:25] + "..." if len(recipient_str) > 25 else recipient_str,
                    'Sí' if email.has_attachments else 'No'
                ])
            
            # Crear tabla
            table = Table(data, colWidths=[0.5*inch, 0.9*inch, 2*inch, 1.2*inch, 1.5*inch, 0.5*inch])
            
            # Estilos de tabla
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ROWHEIGHT', (0, 1), (-1, -1), 18),
            ]))
            
            story.append(table)
        else:
            story.append(Paragraph("No hay correos para mostrar", styles['Normal']))
        
        # Construir PDF
        doc.build(story)
        
        return pdf_buffer.getvalue()
    
    @staticmethod
    def to_excel(emails, include_audit_info=True):
        """
        Exporta correos a formato Excel con información de auditoría
        
        Args:
            emails (list): Lista de objetos Email
            include_audit_info (bool): Si incluir información de auditoría
            
        Returns:
            bytes: Contenido Excel como bytes
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Correos"
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = [
            'ID',
            'Fecha de Envío',
            'Asunto',
            'Remitente',
            'Destinatarios',
            'Importancia',
            'Archivos Adjuntos',
            'Fecha de Creación en Sistema'
        ]
        
        if include_audit_info:
            headers.extend(['Registrado por Usuario', 'Fecha Última Modificación', 'Modificado por'])
        
        # Escribir encabezados
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Importar User una sola vez
        from app.models import User
        
        # Escribir datos
        row_num = 2
        for email in emails:
            try:
                # Parsear destinatarios
                try:
                    recipients = json.loads(email.recipients)
                    recipient_emails = [r['email'] for r in recipients]
                    recipient_str = '; '.join(recipient_emails)
                except:
                    recipient_str = email.recipients
                
                row_data = [
                    email.id,
                    email.sent_date.strftime('%Y-%m-%d %H:%M:%S'),
                    email.subject,
                    email.sender,
                    recipient_str,
                    email.importance,
                    'Sí' if email.has_attachments else 'No',
                    email.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ]
                
                if include_audit_info:
                    try:
                        # Obtener último log de auditoría
                        last_audit = email.audit_logs.order_by(AuditLog.timestamp.desc()).first() if hasattr(email, 'audit_logs') else None
                        if last_audit:
                            creator_user = User.query.filter_by(id=last_audit.user_id).first()
                            row_data.append(creator_user.username if creator_user else "Desconocido")
                            row_data.append(last_audit.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
                            row_data.append(creator_user.email if creator_user else "N/A")
                        else:
                            row_data.extend(["N/A", "N/A", "N/A"])
                    except Exception as audit_err:
                        # Si hay error en auditoría, usar valores por defecto
                        current_app.logger.warning(f"Error obteniendo auditoría para email {email.id}: {str(audit_err)}")
                        row_data.extend(["N/A", "N/A", "N/A"])
                
                # Escribir fila en Excel
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                row_num += 1
                
            except Exception as email_err:
                # Si hay error procesando un email, log y continuar con el siguiente
                current_app.logger.error(f"Error procesando email {email.id}: {str(email_err)}", exc_info=True)
                continue
        
        # Ajustar ancho de columnas
        column_widths = [5, 20, 30, 20, 30, 12, 15, 20]
        if include_audit_info:
            column_widths.extend([18, 20, 20])
        
        for col_num, width in enumerate(column_widths, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = width
        
        # Guardar a bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
